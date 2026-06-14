from __future__ import annotations

import math
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path("aouy.xlsx")
OUTPUT_DIR = Path("outputs")
OUTPUT = OUTPUT_DIR / "true_af_nlp_sentiment_dashboard.xlsx"


TEXT_COL = "ข้อความคอมเมนต์ล่าสุด"

NEGATIVE_TERMS = [
    "ไม่อยาก",
    "ไม่ดู",
    "ไม่ได้ดู",
    "ดูไม่ได้",
    "หาช่องดูไม่ได้",
    "ยาก",
    "ยุ่งยาก",
    "เสียดาย",
    "เซง",
    "เซ็ง",
    "อึน",
    "ไม่เคลียร์",
    "ต้องสมัคร",
    "ต้องใช้",
    "แพ็ก",
    "แพค",
    "เสียตังค์",
    "เสียเงิน",
    "จ่าย",
    "จำกัด",
    "exclusive",
    "ไม่ใช่ลูกค้าทรู",
    "คืนตัง",
    "ไม่ผ่าน",
    "มั่ว",
    "เงียบ",
    "แย่",
    "ด่า",
    "ว่า",
    "บ่น",
    "ไม่ควร",
    "เข้าไม่ถึง",
    "ลำบาก",
]

ACCESS_TERMS = [
    "ดูไม่ได้",
    "ไม่ได้ดู",
    "หาช่องดูไม่ได้",
    "ช่องทาง",
    "รับชม",
    "ถ่ายทอดสด",
    "ฟรีทีวี",
    "แอป",
    "app",
    "youtube",
    "now",
    "รีรัน",
    "ช่อง",
    "คอนเสิร์ต",
    "concert",
    "ดูสด",
]

TRUE_PACKAGE_TERMS = [
    "ทรู",
    "true",
    "dtac",
    "ดีเทค",
    "ซิม",
    "ลูกค้าทรู",
    "แพ็ก",
    "แพค",
    "package",
    "99",
    "199",
    "exclusive",
    "เสียตังค์",
    "เสียเงิน",
    "จ่าย",
    "สมัคร",
]

COMPLAINT_TERMS = [
    "ด่า",
    "แย่",
    "ไม่ผ่าน",
    "มั่ว",
    "เซง",
    "เซ็ง",
    "ไม่เคลียร์",
    "ยุ่งยาก",
    "จำกัด",
    "ไม่ควร",
    "ไม่อยาก",
    "คืนตัง",
    "อึน",
]


def hits(text: str, terms: list[str]) -> list[str]:
    lower = text.lower()
    found: list[str] = []
    for term in terms:
        if term.lower() in lower and term not in found:
            found.append(term)
    return found


def sentiment_bucket(score: int) -> str:
    if score >= 70:
        return "Very Negative"
    if score >= 40:
        return "Negative"
    if score >= 20:
        return "Mixed / Concern"
    return "Neutral / Other"


def classify(text: str) -> dict[str, object]:
    neg = hits(text, NEGATIVE_TERMS)
    access = hits(text, ACCESS_TERMS)
    true_pkg = hits(text, TRUE_PACKAGE_TERMS)
    complaint = hits(text, COMPLAINT_TERMS)
    unable = hits(text, ["ไม่ได้ดู", "ดูไม่ได้", "หาช่องดูไม่ได้", "ดูทางไหน", "ต้องสมัคร", "ไม่ใช่ลูกค้าทรู"])
    score = min(
        100,
        12 * len(neg)
        + 8 * len(access)
        + 10 * len(true_pkg)
        + 10 * len(complaint)
        + (12 if unable else 0),
    )
    return {
        "negative_score": score,
        "sentiment": sentiment_bucket(score),
        "negative_terms": ", ".join(neg),
        "access_terms": ", ".join(access),
        "true_package_terms": ", ".join(true_pkg),
        "complaint_terms": ", ".join(complaint),
        "is_access_complaint": "Yes" if access else "No",
        "is_true_package_issue": "Yes" if true_pkg else "No",
        "is_unable_to_watch": "Yes" if unable else "No",
        "is_direct_complaint_language": "Yes" if complaint else "No",
    }


def split_phrases(text: str) -> list[str]:
    parts = re.split(r"[\n\r.!?。]|ค่ะ|ครับ|นะคะ|นะครับ", text)
    cleaned = []
    for part in parts:
        part = re.sub(r"\s+", " ", part).strip()
        if len(part) >= 12:
            cleaned.append(part)
    return cleaned


def phrase_score(phrase: str) -> int:
    return min(
        100,
        12 * len(hits(phrase, NEGATIVE_TERMS))
        + 8 * len(hits(phrase, ACCESS_TERMS))
        + 10 * len(hits(phrase, TRUE_PACKAGE_TERMS))
        + 10 * len(hits(phrase, COMPLAINT_TERMS)),
    )


def pct(part: int, whole: int) -> float:
    return part / whole if whole else 0


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    headers = list(df.columns)
    for c, header in enumerate(headers, start_col):
        ws.cell(start_row, c, header)
    for r, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c, value in enumerate(row, start_col):
            if pd.isna(value):
                value = None
            ws.cell(r, c, value)


def style_table(ws, min_row: int, max_row: int, min_col: int, max_col: int, header_fill: str = "1F4E5F") -> None:
    thin = Side(style="thin", color="D9E2E7")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Tahoma", size=10, color="1E293B")
    for cell in ws[min_row][min_col - 1 : max_col]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width: int = 55) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, math.ceil(len(value) * 0.9) + 2))
        ws.column_dimensions[letter].width = width


def add_kpi(ws, cell: str, label: str, value: str, fill: str) -> None:
    ws[cell] = label
    ws[cell].font = Font(name="Tahoma", bold=True, color="FFFFFF", size=10)
    ws[cell].fill = PatternFill("solid", fgColor=fill)
    ws[cell].alignment = Alignment(horizontal="center")
    value_cell = ws.cell(ws[cell].row + 1, ws[cell].column)
    value_cell.value = value
    value_cell.font = Font(name="Tahoma", bold=True, color="0F172A", size=18)
    value_cell.fill = PatternFill("solid", fgColor="F8FAFC")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = Border(bottom=Side(style="medium", color=fill))


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    df = pd.read_excel(INPUT, sheet_name="Comments_Data")
    df[TEXT_COL] = df[TEXT_COL].fillna("").astype(str)

    analysis = pd.DataFrame([classify(text) for text in df[TEXT_COL]])
    detail = pd.concat([df, analysis], axis=1)

    total = len(detail)
    counts = detail["sentiment"].value_counts().reindex(
        ["Very Negative", "Negative", "Mixed / Concern", "Neutral / Other"], fill_value=0
    )
    issue_summary = pd.DataFrame(
        [
            ["มีประเด็นช่องทางรับชม / ดูรายการไม่ได้", int((detail["is_access_complaint"] == "Yes").sum())],
            ["เกี่ยวกับซิม/ลูกค้าทรู/แพ็กเกจ/จ่ายเงิน", int((detail["is_true_package_issue"] == "Yes").sum())],
            ["พูดตรง ๆ ว่าดูไม่ได้/ไม่ได้ดู/หาไม่เจอ/ต้องสมัคร", int((detail["is_unable_to_watch"] == "Yes").sum())],
            ["มีถ้อยคำต่อว่าเชิงลบชัดเจน", int((detail["is_direct_complaint_language"] == "Yes").sum())],
        ],
        columns=["Issue", "Comments"],
    )
    issue_summary["Share"] = issue_summary["Comments"].apply(lambda x: pct(int(x), total))

    sentiment_summary = pd.DataFrame(
        {
            "Sentiment": counts.index,
            "Comments": counts.values,
            "Share": [pct(int(v), total) for v in counts.values],
        }
    )

    term_counter: Counter[str] = Counter()
    for term_list in detail["negative_terms"].fillna(""):
        for term in [t.strip() for t in str(term_list).split(",") if t.strip()]:
            term_counter[term] += 1
    keyword_summary = pd.DataFrame(term_counter.most_common(25), columns=["Keyword", "Frequency"])

    phrase_rows = []
    for _, row in detail.iterrows():
        text = str(row[TEXT_COL])
        for phrase in split_phrases(text):
            score = phrase_score(phrase)
            if score >= 18:
                phrase_rows.append(
                    {
                        "negative_phrase_score": score,
                        "sentiment": sentiment_bucket(min(100, score)),
                        "ชื่อผู้แสดงความคิดเห็น": row.get("ชื่อผู้แสดงความคิดเห็น", ""),
                        "เวลาที่แสดงความคิดเห็น": row.get("เวลาที่แสดงความคิดเห็น", ""),
                        "จำนวนกดไลค์และแสดงความรู้สึก": row.get("จำนวนกดไลค์และแสดงความรู้สึก", 0),
                        "คำพูด/วลี negative": phrase[:500],
                        "terms": ", ".join(
                            hits(phrase, NEGATIVE_TERMS + ACCESS_TERMS + TRUE_PACKAGE_TERMS + COMPLAINT_TERMS)
                        ),
                    }
                )
    phrase_df = pd.DataFrame(phrase_rows).sort_values("negative_phrase_score", ascending=False).head(30)

    top_comments = detail.sort_values("negative_score", ascending=False).head(20)[
        [
            "negative_score",
            "sentiment",
            "ชื่อผู้แสดงความคิดเห็น",
            "เวลาที่แสดงความคิดเห็น",
            "is_access_complaint",
            "is_true_package_issue",
            "is_unable_to_watch",
            "complaint_terms",
            TEXT_COL,
        ]
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    title_fill = "143642"
    accent = "D94F30"
    teal = "1F7A7A"
    amber = "E8A317"
    slate = "475569"

    ws.merge_cells("A1:L1")
    ws["A1"] = "True AF NLP Sentiment Dashboard"
    ws["A1"].fill = PatternFill("solid", fgColor=title_fill)
    ws["A1"].font = Font(name="Tahoma", bold=True, color="FFFFFF", size=20)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:L3")
    ws["A2"] = (
        "วิเคราะห์คอมเมนต์จาก aouy.xlsx เพื่อดูระดับ sentiment เชิงลบ ประเด็นดูรายการ AF ไม่ได้ "
        "การต้องใช้ซิม/ลูกค้า/แพ็กเกจ True และคำพูดที่ต่อว่าในเชิง negative"
    )
    ws["A2"].font = Font(name="Tahoma", color="334155", size=11)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")

    negative_total = int(counts["Very Negative"] + counts["Negative"])
    concern_total = int(negative_total + counts["Mixed / Concern"])
    avg_score = round(float(detail["negative_score"].mean()), 1)
    access_count = int((detail["is_access_complaint"] == "Yes").sum())
    true_count = int((detail["is_true_package_issue"] == "Yes").sum())

    add_kpi(ws, "A5", "Total comments", f"{total:,}", slate)
    add_kpi(ws, "C5", "Negative comments", f"{negative_total:,} ({pct(negative_total, total):.0%})", accent)
    add_kpi(ws, "E5", "All concerns", f"{concern_total:,} ({pct(concern_total, total):.0%})", amber)
    add_kpi(ws, "G5", "Access issue", f"{access_count:,} ({pct(access_count, total):.0%})", teal)
    add_kpi(ws, "I5", "True/package issue", f"{true_count:,} ({pct(true_count, total):.0%})", "7C3AED")
    add_kpi(ws, "K5", "Avg negative score", f"{avg_score}", "0F766E")

    start = 9
    ws[f"A{start}"] = "Sentiment distribution"
    ws[f"A{start}"].font = Font(name="Tahoma", bold=True, size=13, color="143642")
    write_df(ws, sentiment_summary, start + 1, 1)
    style_table(ws, start + 1, start + 1 + len(sentiment_summary), 1, 3, title_fill)
    for row in range(start + 2, start + 2 + len(sentiment_summary)):
        ws.cell(row, 3).number_format = "0%"

    chart = DoughnutChart()
    chart.title = "Sentiment Share"
    data = Reference(ws, min_col=2, min_row=start + 1, max_row=start + len(sentiment_summary) + 1)
    labels = Reference(ws, min_col=1, min_row=start + 2, max_row=start + len(sentiment_summary) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.holeSize = 55
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.height = 7
    chart.width = 10
    ws.add_chart(chart, "E9")

    issue_start = 18
    ws[f"A{issue_start}"] = "Key issue summary"
    ws[f"A{issue_start}"].font = Font(name="Tahoma", bold=True, size=13, color="143642")
    write_df(ws, issue_summary, issue_start + 1, 1)
    style_table(ws, issue_start + 1, issue_start + 1 + len(issue_summary), 1, 3, title_fill)
    for row in range(issue_start + 2, issue_start + 2 + len(issue_summary)):
        ws.cell(row, 3).number_format = "0%"

    issue_chart = BarChart()
    issue_chart.title = "Issue Volume"
    issue_chart.type = "bar"
    issue_chart.style = 10
    issue_chart.y_axis.title = "Issue"
    issue_chart.x_axis.title = "Comments"
    data = Reference(ws, min_col=2, min_row=issue_start + 1, max_row=issue_start + len(issue_summary) + 1)
    labels = Reference(ws, min_col=1, min_row=issue_start + 2, max_row=issue_start + len(issue_summary) + 1)
    issue_chart.add_data(data, titles_from_data=True)
    issue_chart.set_categories(labels)
    issue_chart.height = 7
    issue_chart.width = 12
    ws.add_chart(issue_chart, "E18")

    quote_start = 28
    ws[f"A{quote_start}"] = "Top negative phrases / คำพูดที่ต่อว่า"
    ws[f"A{quote_start}"].font = Font(name="Tahoma", bold=True, size=13, color="143642")
    dash_quotes = phrase_df[["negative_phrase_score", "คำพูด/วลี negative", "terms"]].head(8)
    write_df(ws, dash_quotes, quote_start + 1, 1)
    style_table(ws, quote_start + 1, quote_start + 1 + len(dash_quotes), 1, 3, accent)
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 38

    ws["H28"] = "Executive insight"
    ws["H28"].font = Font(name="Tahoma", bold=True, size=13, color="143642")
    insights = [
        f"Negative + Very Negative = {negative_total}/{total} comments ({pct(negative_total, total):.0%}).",
        f"ถ้านับกลุ่ม Mixed / Concern ร่วมด้วย มี {concern_total}/{total} comments ({pct(concern_total, total):.0%}) ที่สะท้อน pain point.",
        f"ประเด็นช่องทางรับชม/ดูไม่ได้พบ {access_count}/{total} comments ({pct(access_count, total):.0%}).",
        f"ประเด็น True/ซิม/แพ็กเกจ/จ่ายเงินพบ {true_count}/{total} comments ({pct(true_count, total):.0%}).",
        "คำที่สะท้อนความไม่พอใจสูง: ดูไม่ได้, ต้องสมัครแพ็ก, ไม่ใช่ลูกค้าทรู, ไม่เคลียร์, ยุ่งยาก, จำกัด, เซ็ง.",
    ]
    for i, text in enumerate(insights, 29):
        ws.merge_cells(start_row=i, start_column=8, end_row=i, end_column=12)
        ws.cell(i, 8, text)
        ws.cell(i, 8).font = Font(name="Tahoma", size=11, color="334155")
        ws.cell(i, 8).alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["H"].width = 24
    for row in range(29, 34):
        ws.row_dimensions[row].height = 38

    for sheet_name, data in [
        ("NLP_Detail", detail),
        ("Negative_Quotes", phrase_df),
        ("Keyword_Summary", keyword_summary),
        ("Top_Negative_Comments", top_comments),
        ("Source_Data", df),
    ]:
        sh = wb.create_sheet(sheet_name)
        sh.sheet_view.showGridLines = False
        write_df(sh, data)
        style_table(sh, 1, len(data) + 1, 1, len(data.columns), title_fill if sheet_name != "Negative_Quotes" else accent)
        sh.freeze_panes = "A2"
        auto_width(sh)
        if sheet_name in {"NLP_Detail", "Top_Negative_Comments", "Negative_Quotes"}:
            sh.conditional_formatting.add(
                "A2:A500",
                ColorScaleRule(start_type="min", start_color="F8FAFC", mid_type="percentile", mid_value=50, mid_color="FDE68A", end_type="max", end_color="F97316"),
            )
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_meta = wb.create_sheet("Method")
    ws_meta.sheet_view.showGridLines = False
    method_rows = pd.DataFrame(
        [
            ["Input", "aouy.xlsx / Comments_Data"],
            ["Text column", TEXT_COL],
            ["Scoring", "Rule-based Thai keyword scoring tailored to access, True package, and complaint language."],
            ["Very Negative", "negative_score >= 70"],
            ["Negative", "negative_score 40-69"],
            ["Mixed / Concern", "negative_score 20-39"],
            ["Neutral / Other", "negative_score < 20"],
            ["Note", "One comment can match multiple issues. The method is transparent and auditable in NLP_Detail."],
        ],
        columns=["Item", "Description"],
    )
    write_df(ws_meta, method_rows)
    style_table(ws_meta, 1, len(method_rows) + 1, 1, 2, title_fill)
    auto_width(ws_meta, 90)

    wb.save(OUTPUT)

    check = load_workbook(OUTPUT)
    required = {"Dashboard", "NLP_Detail", "Negative_Quotes", "Keyword_Summary", "Top_Negative_Comments", "Source_Data", "Method"}
    missing = required.difference(check.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    if not check["Dashboard"]._charts:
        raise RuntimeError("Dashboard charts were not created")
    print(f"Saved {OUTPUT}")
    print(f"Rows analyzed: {total}")
    print(f"Negative+Very Negative: {negative_total} ({pct(negative_total, total):.1%})")
    print(f"All concern: {concern_total} ({pct(concern_total, total):.1%})")
    print(f"Access issue: {access_count} ({pct(access_count, total):.1%})")
    print(f"True/package issue: {true_count} ({pct(true_count, total):.1%})")


if __name__ == "__main__":
    main()
